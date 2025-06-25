"""
Excel Integration Adapter for Pynomaly

This module provides Excel integration capabilities for importing datasets
and exporting anomaly detection results with advanced formatting, charts,
and visualizations.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

try:
    import openpyxl
    from openpyxl.chart import Reference, ScatterChart, Series
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlsxwriter

    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False

from ...application.dto.export_options import ExportOptions
from ...domain.entities.dataset import Dataset
from ...domain.entities.detection_result import DetectionResult
from ...shared.protocols.export_protocol import ExportProtocol
from ...shared.protocols.import_protocol import ImportProtocol

logger = logging.getLogger(__name__)


class ExcelAdapter(ExportProtocol, ImportProtocol):
    """
    Excel adapter for importing datasets and exporting anomaly detection results.

    Supports both openpyxl and xlsxwriter for different use cases:
    - openpyxl: Reading existing Excel files and basic writing
    - xlsxwriter: Advanced formatting, charts, and performance optimization
    """

    def __init__(self):
        """Initialize Excel adapter with dependency checks."""
        if not OPENPYXL_AVAILABLE and not XLSXWRITER_AVAILABLE:
            raise ImportError(
                "Excel adapter requires either openpyxl or xlsxwriter. "
                "Install with: pip install openpyxl xlsxwriter"
            )

        self._openpyxl_available = OPENPYXL_AVAILABLE
        self._xlsxwriter_available = XLSXWRITER_AVAILABLE

        logger.info(
            f"Excel adapter initialized. "
            f"openpyxl: {self._openpyxl_available}, "
            f"xlsxwriter: {self._xlsxwriter_available}"
        )

    def export_results(
        self,
        results: DetectionResult,
        file_path: str | Path,
        options: ExportOptions | None = None,
    ) -> dict[str, Any]:
        """
        Export anomaly detection results to Excel file with formatting and charts.

        Args:
            results: Detection results to export
            file_path: Output Excel file path
            options: Export configuration options

        Returns:
            Dictionary containing export metadata and statistics
        """
        if not self._xlsxwriter_available and not self._openpyxl_available:
            raise RuntimeError("No Excel library available for export")

        file_path = Path(file_path)
        options = options or ExportOptions()

        # Use xlsxwriter for advanced formatting if available
        if self._xlsxwriter_available and options.use_advanced_formatting:
            return self._export_with_xlsxwriter(results, file_path, options)
        elif self._openpyxl_available:
            return self._export_with_openpyxl(results, file_path, options)
        else:
            raise RuntimeError("No suitable Excel library for requested export options")

    def _export_with_xlsxwriter(
        self, results: DetectionResult, file_path: Path, options: ExportOptions
    ) -> dict[str, Any]:
        """Export using xlsxwriter for advanced formatting and charts."""
        import xlsxwriter

        workbook = xlsxwriter.Workbook(str(file_path))

        try:
            # Define formats
            header_format = workbook.add_format(
                {
                    "bold": True,
                    "bg_color": "#4472C4",
                    "font_color": "white",
                    "border": 1,
                    "align": "center",
                    "valign": "vcenter",
                }
            )

            anomaly_format = workbook.add_format(
                {
                    "bg_color": "#FFE6E6",
                    "font_color": "#CC0000",
                    "bold": True,
                    "border": 1,
                }
            )

            normal_format = workbook.add_format({"bg_color": "#E6F3E6", "border": 1})

            score_format = workbook.add_format(
                {"num_format": "0.000", "border": 1, "align": "center"}
            )

            # Create main results worksheet
            self._create_results_worksheet(
                workbook,
                results,
                header_format,
                anomaly_format,
                normal_format,
                score_format,
            )

            # Create summary worksheet
            self._create_summary_worksheet(workbook, results, header_format)

            # Create charts if requested
            if options.include_charts:
                self._create_charts_worksheet(workbook, results)

            # Create metadata worksheet
            self._create_metadata_worksheet(workbook, results, options, header_format)

            workbook.close()

            return {
                "file_path": str(file_path),
                "export_time": datetime.now().isoformat(),
                "total_samples": len(results.scores),
                "anomalies_count": sum(
                    1 for score in results.scores if score.is_anomaly
                ),
                "worksheets": ["Results", "Summary", "Charts", "Metadata"],
            }

        except Exception as e:
            workbook.close()
            raise RuntimeError(f"Failed to export results with xlsxwriter: {e}")

    def _export_with_openpyxl(
        self, results: DetectionResult, file_path: Path, options: ExportOptions
    ) -> dict[str, Any]:
        """Export using openpyxl for basic formatting."""
        import openpyxl

        workbook = openpyxl.Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)

        # Create results worksheet
        results_ws = workbook.create_sheet("Results")
        self._populate_results_openpyxl(results_ws, results)

        # Create summary worksheet
        summary_ws = workbook.create_sheet("Summary")
        self._populate_summary_openpyxl(summary_ws, results)

        # Create metadata worksheet
        metadata_ws = workbook.create_sheet("Metadata")
        self._populate_metadata_openpyxl(metadata_ws, results, options)

        workbook.save(str(file_path))

        return {
            "file_path": str(file_path),
            "export_time": datetime.now().isoformat(),
            "total_samples": len(results.scores),
            "anomalies_count": sum(1 for score in results.scores if score.is_anomaly),
            "worksheets": ["Results", "Summary", "Metadata"],
        }

    def _create_results_worksheet(
        self,
        workbook,
        results: DetectionResult,
        header_format,
        anomaly_format,
        normal_format,
        score_format,
    ):
        """Create the main results worksheet with xlsxwriter."""
        worksheet = workbook.add_worksheet("Results")

        # Headers
        headers = [
            "Index",
            "Anomaly Score",
            "Is Anomaly",
            "Confidence",
            "Feature Values",
        ]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        # Data rows
        for row, (idx, score) in enumerate(
            zip(range(len(results.scores)), results.scores, strict=False), 1
        ):
            # Index
            worksheet.write(row, 0, idx)

            # Anomaly score
            worksheet.write(row, 1, float(score.value), score_format)

            # Is anomaly (get from labels array)
            is_anomaly = bool(results.labels[idx])
            cell_format = anomaly_format if is_anomaly else normal_format
            worksheet.write(row, 2, "YES" if is_anomaly else "NO", cell_format)

            # Confidence
            confidence = getattr(score, "confidence", None)
            if confidence is not None:
                worksheet.write(row, 3, float(confidence), score_format)
            else:
                worksheet.write(row, 3, "N/A")

            # Feature values (if available)
            if hasattr(results, "features") and results.features is not None:
                if idx < len(results.features):
                    feature_str = str(results.features[idx])[
                        :50
                    ]  # Truncate for display
                    worksheet.write(row, 4, feature_str)

        # Auto-adjust column widths
        worksheet.set_column(0, 0, 8)  # Index
        worksheet.set_column(1, 1, 15)  # Score
        worksheet.set_column(2, 2, 12)  # Is Anomaly
        worksheet.set_column(3, 3, 12)  # Confidence
        worksheet.set_column(4, 4, 30)  # Features

    def _create_summary_worksheet(
        self, workbook, results: DetectionResult, header_format
    ):
        """Create summary statistics worksheet."""
        worksheet = workbook.add_worksheet("Summary")

        # Calculate statistics
        scores = [float(score.value) for score in results.scores]
        anomalies = [bool(label) for label in results.labels]

        total_samples = len(scores)
        anomaly_count = sum(anomalies)
        normal_count = total_samples - anomaly_count
        anomaly_rate = (anomaly_count / total_samples) * 100 if total_samples > 0 else 0

        avg_score = sum(scores) / len(scores) if scores else 0
        min_score = min(scores) if scores else 0
        max_score = max(scores) if scores else 0

        # Write summary data
        summary_data = [
            ["Metric", "Value"],
            ["Total Samples", total_samples],
            ["Anomalies Detected", anomaly_count],
            ["Normal Samples", normal_count],
            ["Anomaly Rate (%)", f"{anomaly_rate:.2f}"],
            ["Average Score", f"{avg_score:.4f}"],
            ["Minimum Score", f"{min_score:.4f}"],
            ["Maximum Score", f"{max_score:.4f}"],
            ["Export Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]

        for row, (metric, value) in enumerate(summary_data):
            if row == 0:  # Header row
                worksheet.write(row, 0, metric, header_format)
                worksheet.write(row, 1, value, header_format)
            else:
                worksheet.write(row, 0, metric)
                worksheet.write(row, 1, value)

        worksheet.set_column(0, 0, 20)  # Metric column
        worksheet.set_column(1, 1, 15)  # Value column

    def _create_charts_worksheet(self, workbook, results: DetectionResult):
        """Create charts and visualizations worksheet."""
        worksheet = workbook.add_worksheet("Charts")

        # Create score distribution chart
        chart = workbook.add_chart({"type": "scatter"})

        # Add data series for anomalies and normal points
        scores = [float(score.value) for score in results.scores]
        indices = list(range(len(scores)))

        chart.add_series(
            {
                "name": "Anomaly Scores",
                "categories": indices,
                "values": scores,
                "marker": {"type": "circle", "size": 5},
            }
        )

        chart.set_title({"name": "Anomaly Score Distribution"})
        chart.set_x_axis({"name": "Sample Index"})
        chart.set_y_axis({"name": "Anomaly Score"})
        chart.set_size({"width": 640, "height": 480})

        worksheet.insert_chart("B2", chart)

    def _create_metadata_worksheet(
        self, workbook, results: DetectionResult, options: ExportOptions, header_format
    ):
        """Create metadata worksheet with export information."""
        worksheet = workbook.add_worksheet("Metadata")

        metadata = [
            ["Property", "Value"],
            ["Export Time", datetime.now().isoformat()],
            ["Detector Type", getattr(results, "detector_name", "Unknown")],
            ["Total Samples", len(results.scores)],
            ["Anomaly Threshold", getattr(results, "threshold", "Auto")],
            ["Export Options", str(options.__dict__ if options else {})],
            ["Pynomaly Version", "1.0.0"],  # TODO: Get from package
        ]

        for row, (prop, value) in enumerate(metadata):
            if row == 0:  # Header row
                worksheet.write(row, 0, prop, header_format)
                worksheet.write(row, 1, value, header_format)
            else:
                worksheet.write(row, 0, prop)
                worksheet.write(row, 1, str(value))

        worksheet.set_column(0, 0, 20)  # Property column
        worksheet.set_column(1, 1, 30)  # Value column

    def _populate_results_openpyxl(self, worksheet, results: DetectionResult):
        """Populate results worksheet using openpyxl."""
        # Headers
        headers = ["Index", "Anomaly Score", "Is Anomaly", "Confidence"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )

        # Data
        for row, (idx, score) in enumerate(
            zip(range(len(results.scores)), results.scores, strict=False), 2
        ):
            worksheet.cell(row=row, column=1, value=idx)
            worksheet.cell(row=row, column=2, value=float(score.value))

            is_anomaly = bool(results.labels[idx])
            is_anomaly_cell = worksheet.cell(
                row=row, column=3, value="YES" if is_anomaly else "NO"
            )
            if is_anomaly:
                is_anomaly_cell.fill = PatternFill(
                    start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                )
                is_anomaly_cell.font = Font(color="CC0000", bold=True)

            confidence = getattr(score, "confidence", None)
            worksheet.cell(
                row=row,
                column=4,
                value=float(confidence) if confidence is not None else "N/A",
            )

    def _populate_summary_openpyxl(self, worksheet, results: DetectionResult):
        """Populate summary worksheet using openpyxl."""
        scores = [float(score.value) for score in results.scores]
        anomalies = [bool(label) for label in results.labels]

        summary_data = [
            ["Metric", "Value"],
            ["Total Samples", len(scores)],
            ["Anomalies Detected", sum(anomalies)],
            [
                "Anomaly Rate (%)",
                (
                    f"{(sum(anomalies) / len(anomalies) * 100):.2f}"
                    if anomalies
                    else "0.00"
                ),
            ],
            [
                "Average Score",
                f"{sum(scores) / len(scores):.4f}" if scores else "0.0000",
            ],
        ]

        for row, (metric, value) in enumerate(summary_data, 1):
            worksheet.cell(row=row, column=1, value=metric)
            worksheet.cell(row=row, column=2, value=value)
            if row == 1:  # Header
                worksheet.cell(row=row, column=1).font = Font(bold=True)
                worksheet.cell(row=row, column=2).font = Font(bold=True)

    def _populate_metadata_openpyxl(
        self, worksheet, results: DetectionResult, options: ExportOptions
    ):
        """Populate metadata worksheet using openpyxl."""
        metadata = [
            ["Property", "Value"],
            ["Export Time", datetime.now().isoformat()],
            ["Total Samples", len(results.scores)],
            ["Detector Type", getattr(results, "detector_name", "Unknown")],
        ]

        for row, (prop, value) in enumerate(metadata, 1):
            worksheet.cell(row=row, column=1, value=prop)
            worksheet.cell(row=row, column=2, value=str(value))
            if row == 1:  # Header
                worksheet.cell(row=row, column=1).font = Font(bold=True)
                worksheet.cell(row=row, column=2).font = Font(bold=True)

    def import_dataset(
        self, file_path: str | Path, options: dict[str, Any] | None = None
    ) -> Dataset:
        """
        Import dataset from Excel file.

        Args:
            file_path: Path to Excel file
            options: Import configuration options

        Returns:
            Dataset object containing the imported data
        """
        if not self._openpyxl_available:
            raise RuntimeError("openpyxl is required for Excel import")

        file_path = Path(file_path)
        options = options or {}

        try:
            # Read Excel file with pandas
            df = pd.read_excel(
                file_path,
                sheet_name=options.get("sheet_name", 0),
                header=options.get("header", 0),
                index_col=options.get("index_col", None),
            )

            # Validate and clean data
            df = self._validate_and_clean_data(df, options)

            # Create Dataset object
            dataset = Dataset(
                data=df,
                name=file_path.stem,
                metadata={
                    "source_file": str(file_path),
                    "import_time": datetime.now().isoformat(),
                    "rows": len(df),
                    "columns": len(df.columns),
                    "column_names": list(df.columns),
                },
            )

            logger.info(
                f"Successfully imported dataset from {file_path}: {len(df)} rows, {len(df.columns)} columns"
            )
            return dataset

        except Exception as e:
            logger.error(f"Failed to import dataset from {file_path}: {e}")
            raise RuntimeError(f"Excel import failed: {e}")

    def _validate_and_clean_data(
        self, df: pd.DataFrame, options: dict[str, Any]
    ) -> pd.DataFrame:
        """Validate and clean imported data."""
        # Remove completely empty rows and columns
        df = df.dropna(how="all").dropna(axis=1, how="all")

        # Handle missing values
        fill_method = options.get("fill_missing", "drop")
        if fill_method == "drop":
            df = df.dropna()
        elif fill_method == "forward":
            df = df.fillna(method="ffill")
        elif fill_method == "backward":
            df = df.fillna(method="bfill")
        elif fill_method == "zero":
            df = df.fillna(0)
        elif fill_method == "mean":
            numeric_columns = df.select_dtypes(include=["number"]).columns
            df[numeric_columns] = df[numeric_columns].fillna(df[numeric_columns].mean())

        # Convert data types
        if options.get("auto_convert_types", True):
            df = df.infer_objects()

        # Validate numeric data for anomaly detection
        numeric_columns = df.select_dtypes(include=["number"]).columns
        if len(numeric_columns) == 0:
            raise ValueError("No numeric columns found for anomaly detection")

        return df

    def get_supported_formats(self) -> list[str]:
        """Return list of supported Excel formats."""
        formats = []
        if self._openpyxl_available:
            formats.extend([".xlsx", ".xlsm"])
        if self._xlsxwriter_available:
            formats.extend([".xlsx"])
        return list(set(formats))  # Remove duplicates

    def validate_file(self, file_path: str | Path) -> bool:
        """Validate if file can be processed by this adapter."""
        file_path = Path(file_path)

        # Check file extension
        if file_path.suffix.lower() not in self.get_supported_formats():
            return False

        # Check if file exists and is readable
        if not file_path.exists() or not file_path.is_file():
            return False

        # Try to open with openpyxl if available
        if self._openpyxl_available:
            try:
                import openpyxl

                workbook = openpyxl.load_workbook(file_path, read_only=True)
                workbook.close()
                return True
            except Exception:
                return False

        return True
